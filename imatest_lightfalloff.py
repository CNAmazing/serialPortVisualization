import os

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils import get_column_letter

from imatest_readcsv_tool import *
from typing import Dict, List, Tuple, Union
class imatest_lightfalloff_csv:
    def __init__(self):
        self.read = imatest_readcsv_tool()
        self.config = {
            "start_step": 'B',
            "end_step": 'J',
            "start_row": 43,
            "end_row": 45,
            "Y": [0.3, 0.59, 0.11],
        }

    def lightsource_math(self, txt: str) -> str:
        if "D6" in txt:
            return "D65光源"
        elif "A" in txt:
            return "A光源"
        elif "84" in txt:
            return "TL84光源"
        elif "2800k" in txt:
            return "2800k"
        elif "3000k" in txt:
            return "3000k"
        elif "3500k" in txt:
            return "3500k"
        elif "4000k" in txt:
            return "4000k"
        elif "4150k" in txt:
            return "4150k"
        elif "5000k" in txt:
            return "5000k"
        elif "6500k" in txt:
            return "6500k"
        elif "6750k" in txt:
            return "6750k"
        return "D65光源"

    def analyze_color_data(self, data_step: List[List[float]], file_name: str) -> Tuple[pd.DataFrame, pd.DataFrame]:
        POSITION_MAP = {
            "UL": "左上", "LL": "左下", "UR": "右上", "LR": "右下",
            "L-Ctr": "左中", "R-Ctr": "右中", "T-Ctr": "中上", "B-Ctr": "中下", "Center": "中心"
        }
        TARGET_POSITIONS = ["UL", "LL", "UR", "LR", "Center"]
        Y_WEIGHTS = pd.Series(self.config["Y"], index=["R", "G", "B"])

        extracted_df = pd.DataFrame(data_step,
                                    columns=["Center", "UL", "LL", "UR", "LR", "L-Ctr", "R-Ctr", "T-Ctr", "B-Ctr"],
                                    index=["R", "G", "B"])
        Y_df = pd.DataFrame(extracted_df.T @ Y_WEIGHTS).T
        Y_df.index = ["Y_Value"]

        center_value = Y_df["Center"].iloc[0]
        k_data = {}
        for pos in TARGET_POSITIONS:
            y_val = Y_df[pos].iloc[0]
            if pos == "Center":
                others = [p for p in TARGET_POSITIONS if p != "Center"]
                k_val = (Y_df[others].sum(axis=1) / (len(others) * center_value) * 100).iloc[0]
            else:
                k_val = y_val / center_value * 100
            k_data[POSITION_MAP[pos]] = {"Y": y_val, "K": f"{k_val:.2f}%"}

        for code, name in POSITION_MAP.items():
            if name in k_data:
                k_data[name].update({
                    "R": extracted_df.loc["R", code],
                    "G": extracted_df.loc["G", code],
                    "B": extracted_df.loc["B", code]
                })

        k_df = pd.DataFrame(k_data).T[["R", "G", "B", "Y", "K"]]
        k_df["name"] = file_name
        k_df["光源"] = self.lightsource_math(file_name)

        ratio_df = extracted_df.rename(columns=POSITION_MAP).T
        ratio_df.index.name = "位置"
        ratio_df = ratio_df.T[["左上", "左中", "左下", "中下", "右下", "右中", "右上", "中上", "中心"]]
        ratio_df["name"] = file_name
        ratio_df["光源"] = self.lightsource_math(file_name)
        ratio_df.index.name = "项目"

        return k_df, ratio_df

    def data_groupby(self, k_df: pd.DataFrame, ratio_df: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame]:
        df_center = k_df.loc[k_df.index == '中心', ['光源', 'K']].copy()
        df_center['K'] = df_center['K'].str.rstrip('%').astype(float)
        k_df = df_center.groupby('光源')['K'].mean().reset_index()
        k_df['K'] = k_df['K'].round(2).astype(str) + '%'

        ratio_df = ratio_df.reset_index()
        value_cols = [col for col in ratio_df.columns if col not in ['项目', 'name', '光源']]
        grouped = ratio_df.groupby(['光源', '项目'])[value_cols].mean().reset_index()
        grouped['name'] = '聚合结果'
        df_indexed = grouped.set_index('项目')

        result_rows = []
        for light, group in df_indexed.groupby('光源'):
            group = group.drop(columns=['光源'])
            R = group.loc['R'].drop('name').astype(float)
            G = group.loc['G'].drop('name').astype(float)
            B = group.loc['B'].drop('name').astype(float)

            def append_row(name, series):
                result_rows.append({'光源': light, '指标': name, **series.to_dict()})

            append_row('R', R)
            append_row('G', G)
            append_row('B', B)
            append_row('R/G', R / G)
            append_row('R/B', R / B)
            append_row('B/G', B / G)

            for ratio_name, series in [('R/G', R / G), ('R/B', R / B), ('B/G', B / G)]:
                cols_ex_center = [col for col in R.index if col != '中心']
                min_ratio = series[cols_ex_center].min() / series['中心']
                max_ratio = series[cols_ex_center].max() / series['中心']
                result_rows.append({
                    '光源': light,
                    '左上': f'MIN({ratio_name})', '左中': f"{min_ratio:.6f}",
                    '右下': f'MAX({ratio_name})', '右中': f"{max_ratio:.6f}"
                })

        result_df = pd.DataFrame(result_rows)
        final_cols = ['光源', '指标'] + [col for col in result_df.columns if col not in ['光源', '指标']]
        return k_df, result_df[final_cols]

    def write_csv(self, k_df, ratio_df, grouped_k_df, grouped_ratio_df, file: str):
        with pd.ExcelWriter(file, engine='openpyxl') as writer:
            k_df.to_excel(writer, sheet_name='亮度均匀性详细信息', index=True)
            ratio_df.to_excel(writer, sheet_name='色彩均匀性详细信息', index=True)
            grouped_k_df.to_excel(writer, sheet_name='亮度均匀性准确率', index=False)
            grouped_ratio_df.to_excel(writer, sheet_name='色彩均匀性准确率', index=False)

        wb = load_workbook(file)

        # 边框与对齐格式
        thin = Side(style='thin', color='000000')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

        def format_sheet(ws):
            for row in ws.iter_rows():
                for cell in row:
                    cell.border = border
                    cell.alignment = center_align

            for col in ws.columns:
                max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
                col_letter = get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max_len + 2

        for sheet_name in wb.sheetnames:
            format_sheet(wb[sheet_name])

        wb.save(file)

    def read_csv(self, path: str) -> Tuple[pd.DataFrame, pd.DataFrame]:
        step_data = self.read.extract_csv_data(self.config["start_step"], self.config["end_step"], self.config["start_row"], self.config["end_row"], path)
        return self.analyze_color_data(step_data["data"], os.path.basename(path))

    def run_all(self, paths: List[str], output_file: str):
        k_list, ratio_list = zip(*[self.read_csv(p) for p in paths])
        k_df = pd.concat(k_list, ignore_index=False)
        ratio_df = pd.concat(ratio_list, ignore_index=False)
        grouped_k_df, grouped_ratio_df = self.data_groupby(k_df, ratio_df)
        self.write_csv(k_df, ratio_df, grouped_k_df, grouped_ratio_df, output_file)

    def _find_summary_csv(self, directory):
        for file in os.listdir(directory):
            if "summary.csv" in file:
                return os.path.join(directory, file)
            if "SFR_lwph.csv"  in file:
                return os.path.join(directory, file)
            if "LF_Y.csv"  in file:
                return os.path.join(directory, file)
            if "csv" in file and not "uniformity_dbase" in file:
                return os.path.join(directory, file)
        return None

if __name__ == '__main__':
    path  = r"C:\WorkSpace\serialPortVisualization\data\0821_2"
    output_file = "./均匀性.xls"
    paths= []
    for  i in os.listdir(path):
        if not ".jpg" in i:
            file = imatest_lightfalloff_csv()._find_summary_csv(os.path.join(path, i))
            paths.append(file)
    combined_df_k = pd.DataFrame()
    combined_df_r = pd.DataFrame()
    for file in paths:
        k_list, ratio_list = imatest_lightfalloff_csv().read_csv(file)
        combined_df_k = pd.concat([combined_df_k, k_list], ignore_index=False)
        combined_df_r = pd.concat([combined_df_r, ratio_list], ignore_index=False)
    print(combined_df_k)
    print(combined_df_r)
    grouped_k_df, grouped_ratio_df = imatest_lightfalloff_csv().data_groupby(combined_df_k, combined_df_r)

    imatest_lightfalloff_csv().run_all(paths, output_file)